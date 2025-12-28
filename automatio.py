import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def file_process(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    for row in range(1, sheet.max_row + 1): #get the row number
        print(row)

    for row in range(2, sheet.max_row + 1):  #get the value of 3 column
        cell = sheet.cell(row, 3)
        print(cell.value)

    for row in range(2, sheet.max_row + 1):  #update 3rd column by multiply 0.9
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
    #2-B
    values = Reference(
        sheet,
        min_row=2,
        max_row=sheet.max_row,
        min_col=2,
        max_col=2
    )
    # Categories → Column 2 (D)
    categories = Reference(
        sheet,
        min_row=2,
        max_row=sheet.max_row,
        min_col=4,
        max_col=4
    )
    chart = BarChart()
    chart.add_data(values, titles_from_data=False)
    chart.set_categories(categories)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)